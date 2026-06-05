from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from .models import Produit, Categorie
from .forms import ProduitForm, CategorieForm


@login_required
def liste_produits(request):
    q = request.GET.get('q', '')
    cat = request.GET.get('categorie', '')
    produits = Produit.objects.filter(actif=True)
    if q:
        produits = produits.filter(Q(nom__icontains=q) | Q(reference__icontains=q))
    if cat:
        produits = produits.filter(categorie_id=cat)
    categories = Categorie.objects.all()
    ctx = {'produits': produits, 'categories': categories, 'q': q, 'cat_sel': cat,
           'ruptures': produits.filter(stock_actuel__lte=models_stock_min()).count() if False else 0}
    return render(request, 'produits/liste.html', ctx)


@login_required
def ajouter_produit(request):
    if request.method == 'POST':
        form = ProduitForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'Produit ajouté avec succès.')
            return redirect('produits:liste')
    else:
        form = ProduitForm()
    return render(request, 'produits/form.html', {'form': form, 'titre': 'Ajouter un produit'})


@login_required
def modifier_produit(request, pk):
    produit = get_object_or_404(Produit, pk=pk)
    if request.method == 'POST':
        form = ProduitForm(request.POST, request.FILES, instance=produit)
        if form.is_valid():
            form.save()
            messages.success(request, 'Produit modifié.')
            return redirect('produits:liste')
    else:
        form = ProduitForm(instance=produit)
    return render(request, 'produits/form.html', {'form': form, 'titre': 'Modifier le produit', 'produit': produit})


@login_required
def detail_produit(request, pk):
    produit = get_object_or_404(Produit, pk=pk)
    return render(request, 'produits/detail.html', {'produit': produit})


@login_required
def supprimer_produit(request, pk):
    produit = get_object_or_404(Produit, pk=pk)
    produit.actif = False
    produit.save()
    messages.success(request, 'Produit archivé.')
    return redirect('produits:liste')


@login_required
def exporter_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Produits"
    headers = ['Référence', 'Nom', 'Catégorie', 'Prix Achat', 'Prix Vente', 'Stock', 'Valeur Stock', 'Marge %']
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1a3c5e", end_color="1a3c5e", fill_type="solid")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    produits = Produit.objects.filter(actif=True)
    for row, p in enumerate(produits, 2):
        ws.cell(row=row, column=1, value=p.reference)
        ws.cell(row=row, column=2, value=p.nom)
        ws.cell(row=row, column=3, value=str(p.categorie) if p.categorie else '')
        ws.cell(row=row, column=4, value=int(p.prix_achat))
        ws.cell(row=row, column=5, value=int(p.prix_vente))
        ws.cell(row=row, column=6, value=p.stock_actuel)
        ws.cell(row=row, column=7, value=int(p.valeur_stock))
        ws.cell(row=row, column=8, value=p.marge)
        if p.en_rupture:
            from openpyxl.styles import PatternFill as PF
            fill = PF(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
            for col in range(1, 9):
                ws.cell(row=row, column=col).fill = fill
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="produits.xlsx"'
    wb.save(response)
    return response


@login_required
def regenerer_qr(request, pk):
    produit = get_object_or_404(Produit, pk=pk)
    produit.qr_code.delete(save=False)
    produit.qr_code = None
    produit.generer_qr_code()
    produit.save()
    messages.success(request, 'QR Code régénéré.')
    return redirect('produits:detail', pk=pk)


# Catégories
@login_required
def liste_categories(request):
    categories = Categorie.objects.all()
    return render(request, 'produits/categories.html', {'categories': categories})


@login_required
def ajouter_categorie(request):
    if request.method == 'POST':
        form = CategorieForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Catégorie créée.')
            return redirect('produits:categories')
    else:
        form = CategorieForm()
    return render(request, 'produits/form_categorie.html', {'form': form})

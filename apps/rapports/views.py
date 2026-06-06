from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.db.models import Sum, Count, Avg
from django.db.models.functions import TruncMonth, TruncDay
from django.utils import timezone
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import pandas as pd
import io
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference, LineChart
from datetime import timedelta

from apps.ventes.models import Vente, LigneVente
from apps.produits.models import Produit
from apps.clients.models import Client
from apps.achats.models import CommandeAchat

from . import ontologie


# ─────────────────────────────────────────────────────────────────────────────
# Dashboard rapports (enrichi)
# ─────────────────────────────────────────────────────────────────────────────

@login_required
def dashboard_rapports(request):
    today = timezone.now().date()
    debut = today - timedelta(days=30)

    ventes = Vente.objects.filter(date__date__gte=debut, statut='confirmee')
    total_ventes = ventes.aggregate(t=Sum('total'))['t'] or 0
    nb_ventes = ventes.count()
    top_produits = (LigneVente.objects
                    .filter(vente__date__date__gte=debut)
                    .values('produit__nom')
                    .annotate(qte=Sum('quantite'), ca=Sum('prix_unitaire'))
                    .order_by('-qte')[:5])

    # --- Métriques réseau ---
    G_ventes = ontologie.construire_graphe_ventes(jours=30)
    stats_reseau = ontologie.statistiques_reseau(G_ventes)
    centralite = ontologie.analyser_centralite(G_ventes, top=5)

    ctx = {
        'total_ventes': total_ventes,
        'nb_ventes': nb_ventes,
        'panier_moyen': int(total_ventes / nb_ventes) if nb_ventes else 0,
        'top_produits': list(top_produits),
        'periode': 30,
        'stats_reseau': stats_reseau,
        'top_clients_reseau': centralite.get('top_degree', []),
        'top_produits_reseau': [
            n for n in centralite.get('top_pagerank', [])
            if n.get('type') == 'produit'
        ][:5],
    }
    return render(request, 'rapports/dashboard.html', ctx)


# ─────────────────────────────────────────────────────────────────────────────
# Graphiques matplotlib existants (améliorés)
# ─────────────────────────────────────────────────────────────────────────────

@login_required
def graphique_ventes(request):
    """Retourne un graphique matplotlib en PNG."""
    jours = int(request.GET.get('jours', 30))
    debut = timezone.now() - timedelta(days=jours)

    ventes_par_jour = (Vente.objects
                       .filter(date__gte=debut, statut__in=['confirmee', 'livree'])
                       .annotate(jour=TruncDay('date'))
                       .values('jour')
                       .annotate(total=Sum('total'), nb=Count('id'))
                       .order_by('jour'))

    df = pd.DataFrame(list(ventes_par_jour))

    fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(12, 7))
    fig.patch.set_facecolor('#f8fafc')

    if not df.empty:
        df['jour'] = pd.to_datetime(df['jour'])
        ax1.fill_between(df['jour'], df['total'], alpha=0.3, color='#1a3c5e')
        ax1.plot(df['jour'], df['total'], color='#1a3c5e', linewidth=2, marker='o', markersize=4)
        ax1.set_title('Chiffre d\'affaires journalier (FCFA)', fontweight='bold', color='#1a3c5e')
        ax1.set_facecolor('#f8fafc')
        ax1.xaxis.set_major_formatter(mdates.DateFormatter('%d/%m'))
        ax1.grid(True, alpha=0.3)
        ax1.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f'{int(x):,}'))
        ax1.spines['top'].set_visible(False)
        ax1.spines['right'].set_visible(False)

        ax2.bar(df['jour'], df['nb'], color='#e67e22', alpha=0.8, width=0.8)
        ax2.set_title('Nombre de ventes par jour', fontweight='bold', color='#e67e22')
        ax2.set_facecolor('#f8fafc')
        ax2.xaxis.set_major_formatter(mdates.DateFormatter('%d/%m'))
        ax2.grid(True, alpha=0.3, axis='y')
        ax2.spines['top'].set_visible(False)
        ax2.spines['right'].set_visible(False)
    else:
        for ax in (ax1, ax2):
            ax.text(0.5, 0.5, 'Aucune donnée', ha='center', va='center',
                    transform=ax.transAxes, fontsize=13, color='#999')
            ax.axis('off')

    plt.tight_layout()
    buffer = io.BytesIO()
    plt.savefig(buffer, format='png', dpi=100, bbox_inches='tight', facecolor='#f8fafc')
    plt.close()
    buffer.seek(0)
    return HttpResponse(buffer.read(), content_type='image/png')


@login_required
def graphique_categories(request):
    """Camembert des ventes par catégorie."""
    from apps.produits.models import Categorie
    data = (LigneVente.objects
            .values('produit__categorie__nom')
            .annotate(total=Sum('prix_unitaire'))
            .order_by('-total')[:8])

    labels = [d['produit__categorie__nom'] or 'Sans catégorie' for d in data]
    values = [float(d['total'] or 0) for d in data]

    if not values:
        labels, values = ['Aucune donnée'], [1]

    colors_list = ['#1a3c5e','#2980b9','#e67e22','#27ae60','#8e44ad','#e74c3c','#16a085','#f39c12']
    fig, ax = plt.subplots(figsize=(8, 6))
    fig.patch.set_facecolor('#f8fafc')
    wedges, texts, autotexts = ax.pie(values, labels=labels, autopct='%1.1f%%',
                                       colors=colors_list[:len(labels)], startangle=90)
    ax.set_title('Ventes par catégorie', fontweight='bold', color='#1a3c5e', fontsize=13)
    plt.tight_layout()

    buffer = io.BytesIO()
    plt.savefig(buffer, format='png', dpi=100, bbox_inches='tight', facecolor='#f8fafc')
    plt.close()
    buffer.seek(0)
    return HttpResponse(buffer.read(), content_type='image/png')


# ─────────────────────────────────────────────────────────────────────────────
# Nouvelles vues ontologie / réseau
# ─────────────────────────────────────────────────────────────────────────────

@login_required
def graphique_reseau_ventes(request):
    """PNG : graphe Client → Produit (ventes)."""
    jours = int(request.GET.get('jours', 60))
    G = ontologie.construire_graphe_ventes(jours=jours)
    png = ontologie.graphique_reseau_png(
        G, titre=f"Réseau Client–Produit ({jours} jours)", figsize=(14, 10))
    return HttpResponse(png, content_type='image/png')


@login_required
def graphique_reseau_achats(request):
    """PNG : graphe Fournisseur → Produit (achats)."""
    jours = int(request.GET.get('jours', 90))
    G = ontologie.construire_graphe_achats(jours=jours)
    png = ontologie.graphique_reseau_png(
        G, titre=f"Réseau Fournisseur–Produit ({jours} jours)",
        figsize=(14, 10))
    return HttpResponse(png, content_type='image/png')


@login_required
def graphique_coachat(request):
    """PNG : graphe de co-achat produit–produit."""
    jours = int(request.GET.get('jours', 60))
    G = ontologie.construire_graphe_coachat(jours=jours)
    png = ontologie.graphique_reseau_png(
        G, titre=f"Associations de produits – co-achat ({jours} jours)",
        figsize=(13, 9))
    return HttpResponse(png, content_type='image/png')


@login_required
def graphique_centralite(request):
    """PNG : barres de centralité (clients/produits influents)."""
    jours = int(request.GET.get('jours', 60))
    G = ontologie.construire_graphe_ventes(jours=jours)
    png = ontologie.graphique_centralite_png(
        G, titre=f"Nœuds centraux du réseau de ventes ({jours} jours)")
    return HttpResponse(png, content_type='image/png')


@login_required
def graphique_flux_financiers(request):
    """PNG : flux financiers mensuels ventes vs achats."""
    jours = int(request.GET.get('jours', 90))
    png = ontologie.graphique_flux_financiers_png(jours=jours)
    return HttpResponse(png, content_type='image/png')


@login_required
def api_graphe_json(request):
    """API JSON pour visualisation D3.js interactive."""
    type_graphe = request.GET.get('type', 'ventes')
    jours = int(request.GET.get('jours', 60))

    if type_graphe == 'achats':
        G = ontologie.construire_graphe_achats(jours=jours)
    elif type_graphe == 'coachat':
        G = ontologie.construire_graphe_coachat(jours=jours)
    else:
        G = ontologie.construire_graphe_ventes(jours=jours)

    data = ontologie.graphe_vers_json(G, max_noeuds=80)
    stats = ontologie.statistiques_reseau(G)
    centralite = ontologie.analyser_centralite(G, top=8)
    communautes = ontologie.detecter_communautes(G)

    return JsonResponse({
        'graphe': data,
        'stats': stats,
        'centralite': centralite,
        'communautes': communautes[:5],
    })


@login_required
def dashboard_ontologie(request):
    """Page d'analyse d'ontologie complète."""
    jours = int(request.GET.get('jours', 60))

    G_ventes = ontologie.construire_graphe_ventes(jours=jours)
    G_achats = ontologie.construire_graphe_achats(jours=jours)
    G_coachat = ontologie.construire_graphe_coachat(jours=jours)

    stats_ventes = ontologie.statistiques_reseau(G_ventes)
    stats_achats = ontologie.statistiques_reseau(G_achats)
    stats_coachat = ontologie.statistiques_reseau(G_coachat)

    centralite_ventes = ontologie.analyser_centralite(G_ventes, top=8)
    communautes_coachat = ontologie.detecter_communautes(G_coachat)

    ctx = {
        'jours': jours,
        'stats_ventes': stats_ventes,
        'stats_achats': stats_achats,
        'stats_coachat': stats_coachat,
        'top_clients': [n for n in centralite_ventes.get('top_degree', []) if n['type'] == 'client'][:6],
        'top_produits_pr': [n for n in centralite_ventes.get('top_pagerank', []) if n['type'] == 'produit'][:6],
        'communautes': communautes_coachat[:6],
    }
    return render(request, 'rapports/ontologie.html', ctx)


# ─────────────────────────────────────────────────────────────────────────────
# Rapports existants (inchangés)
# ─────────────────────────────────────────────────────────────────────────────

@login_required
def rapport_excel_complet(request):
    """Rapport Excel multi-feuilles avec graphiques openpyxl + feuille réseau."""
    wb = openpyxl.Workbook()

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1a3c5e', end_color='1a3c5e', fill_type='solid')
    alt_fill = PatternFill(start_color='EBF5FB', end_color='EBF5FB', fill_type='solid')
    center = Alignment(horizontal='center')
    thin = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )

    def style_header(ws, row, cols):
        for col in range(1, cols+1):
            c = ws.cell(row=row, column=col)
            c.font = header_font
            c.fill = header_fill
            c.alignment = center
            c.border = thin

    # Feuille 1 : Résumé
    ws1 = wb.active
    ws1.title = "Résumé"
    today = timezone.now().date()
    debut_mois = today.replace(day=1)

    ventes_mois = Vente.objects.filter(date__date__gte=debut_mois)
    ca_mois = ventes_mois.aggregate(t=Sum('total'))['t'] or 0
    nb_clients = Client.objects.filter(actif=True).count()
    nb_produits = Produit.objects.filter(actif=True).count()
    ruptures = Produit.objects.filter(actif=True, stock_actuel__lte=0).count()

    ws1['A1'] = f'RAPPORT ERP — {today.strftime("%d/%m/%Y")}'
    ws1['A1'].font = Font(size=16, bold=True, color='1a3c5e')
    ws1.merge_cells('A1:D1')

    kpis = [
        ('CA du mois (FCFA)', int(ca_mois)),
        ('Nombre de ventes', ventes_mois.count()),
        ('Clients actifs', nb_clients),
        ('Produits actifs', nb_produits),
        ('Produits en rupture', ruptures),
    ]
    for i, (label, val) in enumerate(kpis, 3):
        ws1.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws1.cell(row=i, column=2, value=val)
    ws1.column_dimensions['A'].width = 30
    ws1.column_dimensions['B'].width = 20

    # Feuille 2 : Ventes du mois
    ws2 = wb.create_sheet("Ventes du mois")
    headers2 = ['N°', 'Date', 'Client', 'Montant (FCFA)', 'Mode paiement', 'Statut']
    for col, h in enumerate(headers2, 1):
        ws2.cell(row=1, column=col, value=h)
    style_header(ws2, 1, len(headers2))

    ventes = Vente.objects.filter(date__date__gte=debut_mois).select_related('client')
    for row, v in enumerate(ventes, 2):
        ws2.cell(row=row, column=1, value=v.numero)
        ws2.cell(row=row, column=2, value=v.date.strftime('%d/%m/%Y'))
        ws2.cell(row=row, column=3, value=v.client.nom if v.client else 'Comptant')
        ws2.cell(row=row, column=4, value=int(v.total))
        ws2.cell(row=row, column=5, value=v.get_mode_paiement_display())
        ws2.cell(row=row, column=6, value=v.get_statut_display())
        if row % 2 == 0:
            for col in range(1, 7):
                ws2.cell(row=row, column=col).fill = alt_fill
    for col in ws2.columns:
        ws2.column_dimensions[col[0].column_letter].width = 18

    # Feuille 3 : Stock
    ws3 = wb.create_sheet("Stock")
    headers3 = ['Référence', 'Produit', 'Catégorie', 'Stock actuel', 'Stock min', 'Valeur stock (FCFA)', 'Alerte']
    for col, h in enumerate(headers3, 1):
        ws3.cell(row=1, column=col, value=h)
    style_header(ws3, 1, len(headers3))

    rouge_fill = PatternFill(start_color='FFE0E0', end_color='FFE0E0', fill_type='solid')
    for row, p in enumerate(Produit.objects.filter(actif=True), 2):
        ws3.cell(row=row, column=1, value=p.reference)
        ws3.cell(row=row, column=2, value=p.nom)
        ws3.cell(row=row, column=3, value=str(p.categorie) if p.categorie else '')
        ws3.cell(row=row, column=4, value=p.stock_actuel)
        ws3.cell(row=row, column=5, value=p.stock_minimum)
        ws3.cell(row=row, column=6, value=int(p.valeur_stock))
        alerte = '⚠️ RUPTURE' if p.en_rupture else 'OK'
        ws3.cell(row=row, column=7, value=alerte)
        if p.en_rupture:
            for col in range(1, 8):
                ws3.cell(row=row, column=col).fill = rouge_fill
    for col in ws3.columns:
        ws3.column_dimensions[col[0].column_letter].width = 20

    # Feuille 4 : Top Produits
    ws4 = wb.create_sheet("Top Produits")
    headers4 = ['Produit', 'Quantité vendue', 'CA (FCFA)']
    for col, h in enumerate(headers4, 1):
        ws4.cell(row=1, column=col, value=h)
    style_header(ws4, 1, 3)

    top = list(LigneVente.objects
               .values('produit__nom')
               .annotate(qte=Sum('quantite'), ca=Sum('prix_unitaire'))
               .order_by('-qte')[:10])
    for row, t in enumerate(top, 2):
        ws4.cell(row=row, column=1, value=t['produit__nom'])
        ws4.cell(row=row, column=2, value=t['qte'])
        ws4.cell(row=row, column=3, value=int(t['ca'] or 0))

    if top:
        chart = BarChart()
        chart.title = "Top 10 Produits (Quantités)"
        chart.style = 10
        data_ref = Reference(ws4, min_col=2, min_row=1, max_row=len(top)+1)
        cats = Reference(ws4, min_col=1, min_row=2, max_row=len(top)+1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats)
        ws4.add_chart(chart, "E2")
    for col in ws4.columns:
        ws4.column_dimensions[col[0].column_letter].width = 25

    # Feuille 5 : Analyse réseau (NetworkX)
    ws5 = wb.create_sheet("Analyse Réseau")
    ws5['A1'] = 'ANALYSE DU RÉSEAU COMMERCIAL (NetworkX)'
    ws5['A1'].font = Font(size=14, bold=True, color='1a3c5e')
    ws5.merge_cells('A1:F1')

    G_ventes = ontologie.construire_graphe_ventes(jours=30)
    stats = ontologie.statistiques_reseau(G_ventes)
    centralite = ontologie.analyser_centralite(G_ventes, top=10)

    ws5['A3'] = 'Métriques du réseau (30 jours)'
    ws5['A3'].font = Font(bold=True, color='FFFFFF')
    ws5['A3'].fill = PatternFill(start_color='1a3c5e', end_color='1a3c5e', fill_type='solid')
    ws5.merge_cells('A3:B3')

    for i, (k, v) in enumerate(stats.items(), 4):
        ws5.cell(row=i, column=1, value=k.replace('_', ' ').title()).font = Font(bold=True)
        ws5.cell(row=i, column=2, value=v)

    row_start = 4 + len(stats) + 2
    ws5.cell(row=row_start, column=1, value='Top Clients (degré entrant)').font = Font(bold=True, color='FFFFFF')
    ws5.cell(row=row_start, column=1).fill = PatternFill(start_color='2980b9', end_color='2980b9', fill_type='solid')
    ws5.cell(row=row_start, column=2, value='Score').font = Font(bold=True, color='FFFFFF')
    ws5.cell(row=row_start, column=2).fill = PatternFill(start_color='2980b9', end_color='2980b9', fill_type='solid')

    for i, n in enumerate(centralite.get('top_degree', []), row_start + 1):
        if n['type'] == 'client':
            ws5.cell(row=i, column=1, value=n['label'])
            ws5.cell(row=i, column=2, value=n['valeur'])
            if i % 2 == 0:
                ws5.cell(row=i, column=1).fill = alt_fill
                ws5.cell(row=i, column=2).fill = alt_fill

    ws5.column_dimensions['A'].width = 35
    ws5.column_dimensions['B'].width = 20

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="rapport_erp_{today}.xlsx"'
    wb.save(response)
    return response


@login_required
def rapport_pandas(request):
    """Analyse avancée avec pandas."""
    ventes = Vente.objects.filter(statut__in=['confirmee', 'livree']).values(
        'numero', 'date', 'total', 'mode_paiement', 'statut'
    )
    df = pd.DataFrame(list(ventes))
    stats = {}
    if not df.empty:
        df['date'] = pd.to_datetime(df['date'])
        df['mois'] = df['date'].dt.to_period('M').astype(str)
        stats = {
            'total_ca': int(df['total'].sum()),
            'moyenne': int(df['total'].mean()),
            'max': int(df['total'].max()),
            'min': int(df['total'].min()),
            'par_mode': df.groupby('mode_paiement')['total'].sum().to_dict(),
            'par_mois': df.groupby('mois')['total'].sum().to_dict(),
            'nb_ventes': len(df),
        }
    return render(request, 'rapports/analyse.html', {'stats': stats})
